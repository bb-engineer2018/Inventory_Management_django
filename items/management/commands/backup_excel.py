
import os
from django.core.management.base import BaseCommand
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

from items.models import Item, ManagementMethod, Order, DeliveredOrder

class Command(BaseCommand):
    help = 'Exports all inventory data to an Excel file.'

    def handle(self, *args, **options):
        self.stdout.write(self.style.SUCCESS('Starting Excel backup...'))

        # バックアップディレクトリのパス
        backup_dir = os.path.join(settings.BASE_DIR, 'backups')
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        file_name = f'物品管理システム_バックアップ_{timestamp}.xlsx'
        file_path = os.path.join(backup_dir, file_name)

        wb = Workbook()

        # 物品マスタシート
        ws_items = wb.active
        ws_items.title = '物品マスタ'
        self._write_header(ws_items, ['ID', '物品名', '納入業者', '規格', '単位', '入数', '保管場所', '管理方法', '納入価格', '削除済み'])
        for item in Item.objects.all():
            ws_items.append([
                item.id, item.name, item.supplier, item.standard, item.unit,
                item.quantity_per_pack, item.storage_location,
                item.management_method.method_name if item.management_method else '',
                item.delivery_price, item.is_deleted
            ])
        self._auto_size_columns(ws_items)

        # 未納品注文シート
        ws_orders = wb.create_sheet(title='未納品注文')
        self._write_header(ws_orders, ['ID', '物品ID', '物品名', '注文日', '注文数'])
        for order in Order.objects.all():
            ws_orders.append([
                order.id, order.item.id, order.item.name, order.order_date.strftime('%Y-%m-%d'), order.order_quantity
            ])
        self._auto_size_columns(ws_orders)

        # 納品済み注文シート
        ws_delivered_orders = wb.create_sheet(title='納品済み注文')
        self._write_header(ws_delivered_orders, ['ID', '物品ID', '物品名', '注文日', '注文数', '納品日'])
        for delivered_order in DeliveredOrder.objects.all():
            ws_delivered_orders.append([
                delivered_order.id, delivered_order.item.id, delivered_order.item.name,
                delivered_order.order_date.strftime('%Y-%m-%d'), delivered_order.order_quantity,
                delivered_order.delivery_date.strftime('%Y-%m-%d')
            ])
        self._auto_size_columns(ws_delivered_orders)

        # 管理方法シート
        ws_methods = wb.create_sheet(title='管理方法')
        self._write_header(ws_methods, ['ID', '管理方法名'])
        for method in ManagementMethod.objects.all():
            ws_methods.append([method.id, method.method_name])
        self._auto_size_columns(ws_methods)

        try:
            wb.save(file_path)
            self.stdout.write(self.style.SUCCESS(f'Excel backup completed successfully: {file_path}'))
        except Exception as e:
            self.stderr.write(self.style.ERROR(f'Error saving Excel file: {e}'))

    def _write_header(self, ws, headers):
        ws.append(headers)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    def _auto_size_columns(self, ws):
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
